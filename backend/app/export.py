"""XLSX export — the breakdown as a live payroll register.

Mirrors the `Sheet8` tab of COMPUTATIONS.xlsx, not `Optimized Structure`. That is a
deliberate choice: `Sheet8` is a **payroll register** — what each person is paid and
what is withheld — which is the document a payroll clerk actually works from.
`Optimized Structure` is a **model**, which answers a different question.

Sheet8's shape, preserved:

    EMPLOYEE | BASIC SALARY | ABSENT/UNDERTIME/LATE | OVERTIME | DE MINIMIS |
    GROSS SALARY | SSS | HDMF | PHIC | GROSS TAXABLE | NON TAXABLE |
    NET TAXABLE INCOME | WITHHOLDING | NET PAY

with deductions carried as negatives and `NET PAY = SUM(NON TAXABLE:WITHHOLDING)`,
exactly as the original does.

**Two deliberate departures, both forced by the model:**

1. A `PRODUCTIVITY INCENTIVE` column is inserted after `DE MINIMIS`. The model has
   three components where Sheet8 had two, and the incentive cannot be folded into
   de minimis: they sit in different exemption tiers under different rules, and
   collapsing them would misstate both the Tier 1 caps and the PHP 90,000 ceiling.
   `NON TAXABLE` is therefore de minimis + incentive.
2. Sheet8's employee rows compute `NET PAY` as `WITHHOLDING/2`, which is half the
   tax rather than net pay — its own example rows use `SUM(K:M)`. The examples are
   right and the employee rows are wrong, so the examples' formula is used here.

**The export carries formulas, not baked numbers.** A spreadsheet of frozen literals
is a screenshot; you cannot change a salary in it and see the effect, which is the
whole reason anyone opens the file. That means the model exists twice — once in
Python, once as Excel formulas — and the divergence risk is handled by
`tests/test_export.py`, which recalculates a generated workbook through Excel and
compares every cell against the engine. Do not weaken that test.

Only Excel-2007-era functions are used (INDEX, MATCH, MIN, MAX, IF, SUM). No
XLOOKUP/FILTER/SORT — they do not survive non-Excel recalculation and truncate
silently, without raising an error.
"""

from datetime import date
from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .engine.models import Result, Scenario
from .engine.tables import SSS_BRACKETS, TRAIN_BRACKETS

# --- house style, lifted from the source workbook --------------------------
FONT = "Arial"
MONEY = "#,##0.00;(#,##0.00);-"
RATE = "0.000"
NAVY = "1F3864"
SLATE = "44546A"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
DERIVED_FILL = PatternFill("solid", fgColor=SLATE)
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")

INPUT_BLUE = "0000FF"      # hardcoded inputs and scenario levers
FORMULA_BLACK = "000000"   # computed on this sheet
CROSSREF_GREEN = "008000"  # points at another sheet

THIN_TOP = Border(top=Side(style="thin"))
THIN_TOPBOT = Border(top=Side(style="thin"), bottom=Side(style="thin"))

REGISTER = "Payroll Register"
REFERENCE = "Reference"
SSS_SHEET = "SSS"

# The register proper — Sheet8's column set, in Sheet8's order.
COLUMNS: List[Tuple[str, int]] = [
    ("EMPLOYEE", 30),
    ("BASIC SALARY", 14),
    ("ABSENT/UNDERTIME/LATE", 15),
    ("OVERTIME", 12),
    ("DE MINIMIS", 13),
    ("PRODUCTIVITY INCENTIVE", 15),
    ("GROSS SALARY", 14),
    ("SSS", 11),
    ("HDMF", 11),
    ("PHIC", 12),
    ("GROSS TAXABLE", 14),
    ("NON TAXABLE", 13),
    ("NET TAXABLE INCOME", 15),
    ("WITHHOLDING", 13),
    ("NET PAY", 14),
]

# Column 16 (P) is left empty as a visual break. The model machinery lives to the
# right of it so the register itself stays exactly Sheet8's shape.
DERIVATION_FIRST_COL = 17
DERIVATION_COLUMNS: List[Tuple[str, int]] = [
    ("RESTRUCTURED?", 13),
    ("13th-mo payment", 14),
    ("Bucket /yr", 13),
    ("Spill /yr", 12),
    ("Annual taxable", 14),
    ("Annual tax", 13),
    ("Base basic /mo", 13),
    ("Base PHIC", 12),
    ("Base ann taxable", 14),
    ("Base annual tax", 14),
    ("TAX SAVED /yr", 14),
]

# Register column letters, for readability in the formula builders below.
BASIC, ABSENT, OVERTIME, DEMINIMIS = "B", "C", "D", "E"
INCENTIVE, GROSS, SSS_C, HDMF_C, PHIC = "F", "G", "H", "I", "J"
GROSS_TAX, NON_TAX, NET_TAX, WTAX, NET_PAY = "K", "L", "M", "N", "O"

RESTR, THIRTEENTH, BUCKET, SPILL = "Q", "R", "S", "T"
ANN_TAXABLE, ANN_TAX = "U", "V"
BASE_BASIC, BASE_PHIC, BASE_TAXABLE, BASE_TAX, SAVED = "W", "X", "Y", "Z", "AA"


def _f(size: int = 10, **kw) -> Font:
    return Font(name=FONT, size=size, **kw)


def _title(ws: Worksheet, row: int, text: str, size: int = 12) -> None:
    ws.cell(row=row, column=1, value=text).font = _f(size, bold=True, color=NAVY)


def _money(cell, value, color: str = FORMULA_BLACK, bold: bool = False):
    cell.value = value
    cell.font = _f(bold=bold, color=color)
    cell.number_format = MONEY
    return cell


def build_workbook(scenario: Scenario, result: Result) -> BytesIO:
    """Render the scenario and its computed result as a live .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = REGISTER
    reference = wb.create_sheet(REFERENCE)
    sss = wb.create_sheet(SSS_SHEET)

    _write_sss_sheet(sss)
    refs = _write_reference_sheet(reference, scenario)
    _write_register(ws, scenario, refs)

    # Excel computes formulas on open; openpyxl writes no cached values, so without
    # this flag some viewers show blanks until a cell is touched.
    wb.calculation.fullCalcOnLoad = True

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_sss_sheet(ws: Worksheet) -> None:
    """SSS contribution schedule — the lookup target, mirroring Sheet9."""
    _title(ws, 1, "SSS CONTRIBUTION SCHEDULE (2025) — employee share")
    for col, head in enumerate(["MSC lower bound", "Employee share"], start=1):
        ws.cell(row=2, column=col, value=head).font = _f(bold=True)

    for i, (lower, share) in enumerate(SSS_BRACKETS):
        r = 3 + i
        _money(ws.cell(row=r, column=1), float(lower), INPUT_BLUE)
        _money(ws.cell(row=r, column=2), float(share), INPUT_BLUE)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.cell(
        row=4 + len(SSS_BRACKETS),
        column=1,
        value=(
            "Matched on the lower bound only. The source workbook's upper-bound column "
            "carries two known data errors; this column is clean and monotonic."
        ),
    ).font = _f(8, italic=True)
    ws.freeze_panes = "A3"


def _write_reference_sheet(ws: Worksheet, scenario: Scenario) -> Dict:
    """De minimis schedule, parameters and the tax table.

    Kept off the register sheet so the register stays exactly Sheet8's shape rather
    than sharing columns with unrelated blocks, the way the original does.
    """
    p = scenario.parameters

    _title(ws, 1, "REFERENCE — schedules, parameters and rates", 14)
    ws.cell(
        row=2,
        column=1,
        value=(
            "Blue cells are inputs; edit them and the register recalculates. "
            "Yellow cells are unverified assumptions."
        ),
    ).font = _f(9, italic=True, color="808080")

    # --- de minimis schedule ---
    dm_title = 4
    _title(ws, dm_title, "1.  DE MINIMIS SCHEDULE  (Tier 1 — does NOT consume the PHP 90,000 ceiling)")
    for col, head in enumerate(
        ["Item", "Statutory cap /mo", "Granted /mo", "Annual", "Basis"], start=1
    ):
        ws.cell(row=dm_title + 1, column=col, value=head).font = _f(9, bold=True)

    first_item = dm_title + 2
    for i, item in enumerate(scenario.deminimis_items):
        r = first_item + i
        ws.cell(row=r, column=1, value=item.label).font = _f()
        _money(ws.cell(row=r, column=2), float(item.statutory_cap_monthly))
        _money(ws.cell(row=r, column=3), float(item.granted_monthly), INPUT_BLUE)
        _money(ws.cell(row=r, column=4), f"=C{r}*12")
        ws.cell(row=r, column=5, value=item.authority).font = _f(8, color="808080")

    last_item = first_item + len(scenario.deminimis_items) - 1
    dm_total = last_item + 1
    ws.cell(row=dm_total, column=1, value="TOTAL DE MINIMIS").font = _f(bold=True)
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        c = _money(
            ws.cell(row=dm_total, column=col),
            f"=SUM({letter}{first_item}:{letter}{last_item})",
            bold=True,
        )
        c.border = THIN_TOPBOT

    # --- parameters ---
    par_title = dm_total + 2
    _title(ws, par_title, "2.  PARAMETERS")
    params = [
        ("PhilHealth employee rate", p.philhealth_rate, RATE, False,
         "5% premium shared equally."),
        ("PhilHealth salary floor", p.philhealth_floor, MONEY, False, ""),
        ("PhilHealth salary ceiling", p.philhealth_ceiling, MONEY, False, ""),
        ("Pag-IBIG employee share", p.pagibig_employee, MONEY, False,
         "2% of the PHP 10,000 fund salary ceiling."),
        ("Benefits exclusion ceiling", p.benefits_ceiling, MONEY, False,
         "NIRC Sec.32(B)(7)(e) — 13th month AND other benefits, combined, per year."),
        ("Cash anchor — carve-out /mo", p.cash_anchor, MONEY, False,
         "Fixes total annual cash, and defines hold-harmless for non-restructured staff."),
        ("Award — baseline comparison only", p.baseline_award, MONEY, False,
         "Feeds the 'before' columns only. Not read by the live calculation."),
        ("Minimum wage — daily", p.min_wage_daily, MONEY, True,
         "ASSUMPTION — verify against the wage order in force."),
        ("Working days divisor", p.working_days, "0", True,
         "ASSUMPTION — 5-day week. Use 313 for a 6-day week."),
    ]
    param_rows: Dict[str, int] = {}
    for i, (name, value, fmt, assumption, note) in enumerate(params):
        r = par_title + 1 + i
        ws.cell(row=r, column=1, value=name).font = _f()
        c = ws.cell(row=r, column=2, value=float(value))
        c.font = _f(color=INPUT_BLUE)
        c.number_format = fmt
        if assumption:
            c.fill = ASSUMPTION_FILL
        if note:
            ws.cell(row=r, column=3, value=note).font = _f(8, italic=True, color="808080")
        param_rows[name] = r

    floor_row = par_title + 1 + len(params)
    ws.cell(row=floor_row, column=1, value="Minimum basic salary /mo (floor)").font = _f()
    _money(
        ws.cell(row=floor_row, column=2),
        f"=B{param_rows['Minimum wage — daily']}*B{param_rows['Working days divisor']}/12",
    )
    ws.cell(row=floor_row, column=3, value="Derived — do not edit.").font = _f(
        8, italic=True, color="808080"
    )

    # --- tax table ---
    tax_title = floor_row + 2
    _title(ws, tax_title, "3.  ANNUAL INCOME TAX TABLE  (TRAIN, effective 2023 onward)")
    for col, head in enumerate(
        ["Over", "Base tax", "Rate on excess", "Threshold"], start=1
    ):
        ws.cell(row=tax_title + 1, column=col, value=head).font = _f(9, bold=True)

    tax_first = tax_title + 2
    for i, (lower, base, rate) in enumerate(TRAIN_BRACKETS):
        r = tax_first + i
        _money(ws.cell(row=r, column=1), float(lower), INPUT_BLUE)
        _money(ws.cell(row=r, column=2), float(base), INPUT_BLUE)
        c = ws.cell(row=r, column=3, value=float(rate))
        c.font = _f(color=INPUT_BLUE)
        c.number_format = "0.0%"
        _money(ws.cell(row=r, column=4), float(lower), INPUT_BLUE)
    tax_last = tax_first + len(TRAIN_BRACKETS) - 1

    for letter, width in [("A", 34), ("B", 16), ("C", 16), ("D", 14), ("E", 52)]:
        ws.column_dimensions[letter].width = width

    return {
        "dm_total": f"{REFERENCE}!$C${dm_total}",
        "ph_rate": f"{REFERENCE}!$B${param_rows['PhilHealth employee rate']}",
        "ph_floor": f"{REFERENCE}!$B${param_rows['PhilHealth salary floor']}",
        "ph_ceiling": f"{REFERENCE}!$B${param_rows['PhilHealth salary ceiling']}",
        "hdmf": f"{REFERENCE}!$B${param_rows['Pag-IBIG employee share']}",
        "ceiling": f"{REFERENCE}!$B${param_rows['Benefits exclusion ceiling']}",
        "anchor": f"{REFERENCE}!$B${param_rows['Cash anchor — carve-out /mo']}",
        "award": f"{REFERENCE}!$B${param_rows['Award — baseline comparison only']}",
        "min_basic": f"{REFERENCE}!$B${floor_row}",
        "tax_first": tax_first,
        "tax_last": tax_last,
    }


def _tax_on(cell: str, refs: Dict) -> str:
    """TRAIN lookup against the Reference sheet. INDEX/MATCH, never XLOOKUP."""
    tf, tl = refs["tax_first"], refs["tax_last"]
    over = f"{REFERENCE}!$A${tf}:$A${tl}"
    m = f"MATCH({cell},{over},1)"
    return (
        f"=INDEX({REFERENCE}!$B${tf}:$B${tl},{m})"
        f"+({cell}-INDEX({REFERENCE}!$D${tf}:$D${tl},{m}))"
        f"*INDEX({REFERENCE}!$C${tf}:$C${tl},{m})"
    )


def _write_register(ws: Worksheet, scenario: Scenario, refs: Dict) -> None:
    """The payroll register, plus the withholding derivation to its right."""
    dm = refs["dm_total"]
    anchor, ceiling, award = refs["anchor"], refs["ceiling"], refs["award"]
    ph_rate, ph_floor, ph_ceiling = refs["ph_rate"], refs["ph_floor"], refs["ph_ceiling"]
    hdmf, min_basic = refs["hdmf"], refs["min_basic"]
    sss_lo, sss_hi = 3, 2 + len(SSS_BRACKETS)

    _title(ws, 1, "PAYROLL REGISTER", 14)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Generated {date.today():%d %B %Y} from the payroll app. Live formulas — "
            "edit any blue cell and the register recalculates. Deductions are shown as "
            "negatives, as in the source workbook."
        ),
    ).font = _f(9, italic=True, color="808080")
    ws.cell(
        row=3,
        column=1,
        value=(
            "Columns A-O are the register. Columns Q-AA derive the withholding from each "
            "employee's annual position and are shown so the figure can be checked, not "
            "taken on trust."
        ),
    ).font = _f(9)

    head_row = 5
    for col, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=head_row, column=col, value=title)
        c.font = _f(9, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    for i, (title, width) in enumerate(DERIVATION_COLUMNS):
        col = DERIVATION_FIRST_COL + i
        c = ws.cell(row=head_row, column=col, value=title)
        c.font = _f(9, bold=True, color="FFFFFF")
        c.fill = DERIVED_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.column_dimensions[get_column_letter(len(COLUMNS) + 1)].width = 3
    ws.row_dimensions[head_row].height = 32

    first = head_row + 1
    for i, employee in enumerate(scenario.employees):
        r = first + i
        ws.cell(row=r, column=1, value=employee.name).font = _f()

        # --- register -----------------------------------------------------
        # Basic is the residual, exactly as Sheet8 has it (B = F - E), extended
        # for the incentive.
        _money(ws.cell(row=r, column=2), f"={GROSS}{r}-{DEMINIMIS}{r}-{INCENTIVE}{r}")
        _money(ws.cell(row=r, column=3), 0, INPUT_BLUE)  # absences, enter negative
        _money(ws.cell(row=r, column=4), 0, INPUT_BLUE)  # overtime
        _money(ws.cell(row=r, column=5), f"={dm}", CROSSREF_GREEN)
        _money(
            ws.cell(row=r, column=6),
            f'=IF({RESTR}{r}="No",{anchor}-{DEMINIMIS}{r},'
            f"MIN(MAX({GROSS}{r}-{DEMINIMIS}{r}-{THIRTEENTH}{r},"
            f"({ceiling}-{THIRTEENTH}{r})/12,0),"
            f"MAX(0,{GROSS}{r}-{DEMINIMIS}{r}-{min_basic})))",
        )
        _money(ws.cell(row=r, column=7), float(employee.signed_gross_monthly), INPUT_BLUE)

        _money(
            ws.cell(row=r, column=8),
            f"=-INDEX({SSS_SHEET}!$B${sss_lo}:$B${sss_hi},"
            f"MATCH({GROSS}{r}-{DEMINIMIS}{r},{SSS_SHEET}!$A${sss_lo}:$A${sss_hi},1))",
            CROSSREF_GREEN,
        )
        _money(ws.cell(row=r, column=9), f"=-{hdmf}", CROSSREF_GREEN)
        # PhilHealth follows Sheet8 in netting absences off the premium base.
        _money(
            ws.cell(row=r, column=10),
            f"=-MIN(MAX({BASIC}{r}+{ABSENT}{r},{ph_floor}),{ph_ceiling})*{ph_rate}",
        )

        _money(
            ws.cell(row=r, column=11),
            f"={GROSS}{r}+{ABSENT}{r}+{OVERTIME}{r}+{SSS_C}{r}+{HDMF_C}{r}+{PHIC}{r}",
        )
        _money(ws.cell(row=r, column=12), f"={DEMINIMIS}{r}+{INCENTIVE}{r}")
        _money(ws.cell(row=r, column=13), f"={GROSS_TAX}{r}-{NON_TAX}{r}")
        _money(ws.cell(row=r, column=14), f"=-{ANN_TAX}{r}/12")
        # Sheet8's own example rows: NET PAY = non-taxable + net taxable + withholding.
        _money(ws.cell(row=r, column=15), f"=SUM({NON_TAX}{r}:{WTAX}{r})", bold=True)

        # --- withholding derivation ---------------------------------------
        c = ws.cell(row=r, column=17, value="Yes" if employee.restructure else "No")
        c.font = _f(color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="center")

        _money(ws.cell(row=r, column=18), f"={GROSS}{r}-{anchor}")
        _money(ws.cell(row=r, column=19), f"={INCENTIVE}{r}*12+{THIRTEENTH}{r}")
        _money(ws.cell(row=r, column=20), f"=MAX(0,{BUCKET}{r}-{ceiling})")
        _money(ws.cell(row=r, column=21), f"={NET_TAX}{r}*12+{SPILL}{r}")
        _money(ws.cell(row=r, column=22), _tax_on(f"{ANN_TAXABLE}{r}", refs))

        _money(ws.cell(row=r, column=23), f"={GROSS}{r}-{anchor}")
        _money(
            ws.cell(row=r, column=24),
            f"=MIN(MAX({BASE_BASIC}{r},{ph_floor}),{ph_ceiling})*{ph_rate}",
        )
        # SSS and HDMF are already negative in the register, so they add here.
        _money(
            ws.cell(row=r, column=25),
            f"=({BASE_BASIC}{r}+{SSS_C}{r}-{BASE_PHIC}{r}+{HDMF_C}{r})*12"
            f"+MAX(0,{BASE_BASIC}{r}+{award}*12-{ceiling})",
        )
        _money(ws.cell(row=r, column=26), _tax_on(f"{BASE_TAXABLE}{r}", refs))
        _money(ws.cell(row=r, column=27), f"={BASE_TAX}{r}-{ANN_TAX}{r}", bold=True)

    last = first + len(scenario.employees) - 1
    total = last + 1
    ws.cell(row=total, column=1, value="TOTAL").font = _f(bold=True)
    ws.cell(row=total, column=1).border = THIN_TOP

    summable = list(range(2, len(COLUMNS) + 1)) + list(
        range(DERIVATION_FIRST_COL + 1, DERIVATION_FIRST_COL + len(DERIVATION_COLUMNS))
    )
    for col in summable:
        letter = get_column_letter(col)
        c = _money(
            ws.cell(row=total, column=col),
            f"=SUM({letter}{first}:{letter}{last})",
            bold=True,
        )
        c.border = THIN_TOP

    notes = [
        "Layout follows the Sheet8 payroll register of COMPUTATIONS.xlsx. Deductions are negatives, and NET PAY = NON TAXABLE + NET TAXABLE INCOME + WITHHOLDING, as in that sheet's example rows.",
        "PRODUCTIVITY INCENTIVE is an added column. The model has three pay components where Sheet8 had two, and the incentive cannot be folded into de minimis: they are different exemption tiers under different rules. NON TAXABLE is de minimis plus the incentive.",
        "ABSENT/UNDERTIME/LATE and OVERTIME are blank inputs for the payroll run. Enter absences as negatives. Both flow into GROSS TAXABLE, and absences also reduce the PhilHealth base.",
        'The incentive column is the optimizer. Employees marked "No" in RESTRUCTURED? receive cash anchor minus de minimis, which pins their taxable basic to the old structure so a change to the de minimis schedule cannot push them into tax.',
        "The productivity incentive requires a threshold that could genuinely fail and a monthly determination recorded per employee. A fixed amount paid unconditionally is regular compensation regardless of the column header.",
        "SSS is looked up on gross less de minimis. PhilHealth is on basic salary only, correctly excluding de minimis and the incentive.",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=total + 2 + i, column=1, value=f"- {note}").font = _f(8)

    # Columns only. Freezing rows as well made the source workbook unscrollable.
    ws.freeze_panes = "B1"


def export_filename() -> str:
    return f"Payroll-Register-{date.today():%Y-%m-%d}.xlsx"
