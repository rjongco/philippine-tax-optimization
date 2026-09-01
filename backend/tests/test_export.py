"""The XLSX export, and the guard against two implementations drifting.

The export carries live Excel formulas so the recipient can edit a salary and watch
the register recalculate. That means the model exists twice — once in Python, once as
formulas. `test_excel_formulas_agree_with_the_engine` recalculates a generated
workbook through Excel and compares every cell to the engine. If the two ever
disagree, that test fails. Do not weaken it.

The recalculation tests need Excel and are skipped where it is unavailable. The
structural tests run everywhere.
"""

import shutil
import subprocess
import sys
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.defaults import default_scenario
from app.engine import compute
from app.export import (
    COLUMNS,
    DERIVATION_COLUMNS,
    DERIVATION_FIRST_COL,
    REFERENCE,
    REGISTER,
    SSS_SHEET,
    build_workbook,
    export_filename,
)

TOLERANCE = Decimal("0.01")

# Register and derivation columns carrying the engine's value unchanged.
DIRECT = {
    2: "basic_monthly",
    5: "deminimis_monthly",
    6: "incentive_monthly",
    7: "signed_gross_monthly",
    13: "net_taxable_monthly",
    15: "net_pay_monthly",
    18: "thirteenth_month_payment",
    19: "bucket_annual",
    20: "spill_annual",
    21: "annual_taxable",
    22: "annual_tax",
    23: "baseline_basic_monthly",
    25: "baseline_annual_taxable",
    26: "baseline_annual_tax",
    27: "tax_saved_annual",
}

# Deductions are carried as negatives in the register, per Sheet8.
NEGATED = {
    8: "sss_employee",
    9: "pagibig_employee",
    10: "philhealth_employee",
    14: "withholding_monthly",
}

# Columns the engine has no single field for.
DERIVED = {
    11: lambda b: (  # GROSS TAXABLE, before removing the non-taxable portion
        b.signed_gross_monthly
        - b.sss_employee
        - b.pagibig_employee
        - b.philhealth_employee
    ),
    12: lambda b: b.deminimis_monthly + b.incentive_monthly,  # NON TAXABLE
    24: lambda b: b.baseline_annual_taxable * 0 + _baseline_phic(b),  # Base PHIC
}


def _baseline_phic(b):
    """PhilHealth on baseline basic — 2.5%, floored at 10k, capped at 100k."""
    base = min(max(b.baseline_basic_monthly, Decimal("10000")), Decimal("100000"))
    return base * Decimal("0.025")


# Every cell in a data row that must be a formula rather than a frozen literal.
FORMULA_COLUMNS = [2, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15] + list(range(18, 28))


@pytest.fixture(scope="module")
def scenario():
    return default_scenario()


@pytest.fixture(scope="module")
def result(scenario):
    return compute(scenario)


@pytest.fixture(scope="module")
def workbook_bytes(scenario, result):
    return build_workbook(scenario, result).getvalue()


def _load(workbook_bytes, name: str, data_only: bool = False):
    path = Path(tempfile.mkdtemp()) / "register.xlsx"
    path.write_bytes(workbook_bytes)
    return openpyxl.load_workbook(path, data_only=data_only)[name]


@pytest.fixture(scope="module")
def sheet(workbook_bytes):
    return _load(workbook_bytes, REGISTER)


def _header_row(ws) -> int:
    for r in range(1, 40):
        if ws.cell(row=r, column=1).value == "EMPLOYEE":
            return r
    raise AssertionError("register header not found")


# --- structure -------------------------------------------------------------


def test_workbook_sheets(workbook_bytes):
    path = Path(tempfile.mkdtemp()) / "s.xlsx"
    path.write_bytes(workbook_bytes)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [REGISTER, REFERENCE, SSS_SHEET]


def test_register_reproduces_sheet8_column_set(sheet):
    """The whole point of this export: Sheet8's shape, not the model's."""
    head = _header_row(sheet)
    expected = [
        "EMPLOYEE", "BASIC SALARY", "ABSENT/UNDERTIME/LATE", "OVERTIME",
        "DE MINIMIS", "PRODUCTIVITY INCENTIVE", "GROSS SALARY", "SSS", "HDMF",
        "PHIC", "GROSS TAXABLE", "NON TAXABLE", "NET TAXABLE INCOME",
        "WITHHOLDING", "NET PAY",
    ]
    actual = [sheet.cell(row=head, column=c).value for c in range(1, len(expected) + 1)]
    assert actual == expected
    assert [t for t, _ in COLUMNS] == expected


def test_derivation_block_is_separated_from_the_register(sheet):
    head = _header_row(sheet)
    # column 16 is the visual break
    assert sheet.cell(row=head, column=len(COLUMNS) + 1).value is None
    for i, (title, _w) in enumerate(DERIVATION_COLUMNS):
        assert sheet.cell(row=head, column=DERIVATION_FIRST_COL + i).value == title


def test_every_employee_has_a_row(sheet, scenario):
    head = _header_row(sheet)
    names = [
        sheet.cell(row=head + 1 + i, column=1).value
        for i in range(len(scenario.employees))
    ]
    assert names == [e.name for e in scenario.employees]
    assert sheet.cell(row=head + 1 + len(scenario.employees), column=1).value == "TOTAL"


def test_computed_cells_are_formulas_not_baked_values(sheet, scenario):
    """A frozen literal is a screenshot. Everything computed must be a formula."""
    head = _header_row(sheet)
    for i in range(len(scenario.employees)):
        r = head + 1 + i
        for col in FORMULA_COLUMNS:
            value = sheet.cell(row=r, column=col).value
            assert isinstance(value, str) and value.startswith("="), (
                f"row {r} column {col} is {value!r}, expected a formula"
            )


def test_gross_salary_is_an_editable_input(sheet, scenario):
    head = _header_row(sheet)
    for i, employee in enumerate(scenario.employees):
        cell = sheet.cell(row=head + 1 + i, column=7)
        assert not isinstance(cell.value, str)
        assert Decimal(str(cell.value)) == employee.signed_gross_monthly
        assert cell.font.color.rgb.endswith("0000FF"), "inputs must be blue"


def test_absence_and_overtime_start_blank_and_editable(sheet, scenario):
    """The register has to be usable for an actual payroll run."""
    head = _header_row(sheet)
    for i in range(len(scenario.employees)):
        r = head + 1 + i
        for col in (3, 4):
            cell = sheet.cell(row=r, column=col)
            assert cell.value == 0
            assert cell.font.color.rgb.endswith("0000FF")


def test_restructure_flag_round_trips(sheet, scenario):
    head = _header_row(sheet)
    for i, employee in enumerate(scenario.employees):
        cell = sheet.cell(row=head + 1 + i, column=DERIVATION_FIRST_COL)
        assert cell.value == ("Yes" if employee.restructure else "No")


def test_no_spilling_array_functions(workbook_bytes):
    """XLOOKUP/FILTER/SORT truncate silently outside Excel. INDEX/MATCH only."""
    banned = ("XLOOKUP", "XMATCH", "FILTER(", "SORT(", "UNIQUE(", "SEQUENCE(")
    path = Path(tempfile.mkdtemp()) / "s.xlsx"
    path.write_bytes(workbook_bytes)
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    upper = cell.value.upper()
                    for token in banned:
                        assert token not in upper, f"{ws.title}!{cell.coordinate}"


def test_full_recalculation_is_requested_on_open(workbook_bytes):
    path = Path(tempfile.mkdtemp()) / "s.xlsx"
    path.write_bytes(workbook_bytes)
    wb = openpyxl.load_workbook(path)
    assert wb.calculation.fullCalcOnLoad is True


def test_filename_names_the_register_and_the_date():
    name = export_filename()
    assert name.startswith("Payroll-Register-") and name.endswith(".xlsx")


# --- the cross-check -------------------------------------------------------


def _recalculate_with_excel(path: Path) -> bool:
    """Open in Excel, rebuild every formula, save. False if Excel is unavailable.

    Uses PowerShell rather than pywin32 so the suite needs no extra dependency.
    LibreOffice (which the xlsx skill's recalc.py drives) is not installed here.
    """
    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        return False

    script = f"""
$ErrorActionPreference = 'Stop'
try {{
  $excel = New-Object -ComObject Excel.Application
}} catch {{ Write-Output 'NO_EXCEL'; exit 0 }}
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open('{path}')
  $excel.CalculateFullRebuild()
  $wb.Save()
  $wb.Close($false)
  Write-Output 'OK'
}} catch {{
  Write-Output "FAILED: $($_.Exception.Message)"
}} finally {{
  $excel.Quit()
  [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
}}
"""
    done = subprocess.run(
        [powershell, "-NoProfile", "-NonInteractive", "-Command", script],
        capture_output=True,
        text=True,
        timeout=180,
    )
    output = done.stdout.strip()
    if "NO_EXCEL" in output:
        return False
    assert "OK" in output, f"Excel recalculation failed: {output}\n{done.stderr}"
    return True


@pytest.fixture(scope="module")
def recalculated(workbook_bytes):
    """A workbook Excel has actually computed. None if Excel is unavailable."""
    path = Path(tempfile.mkdtemp()) / "recalc.xlsx"
    path.write_bytes(workbook_bytes)
    if sys.platform != "win32" or not _recalculate_with_excel(path):
        return None
    return openpyxl.load_workbook(path, data_only=True)


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM is Windows-only")
def test_no_formula_errors_after_recalculation(recalculated):
    if recalculated is None:
        pytest.skip("Excel not available on this machine")

    errors = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")
    bad = [
        f"{ws.title}!{c.coordinate}={c.value}"
        for ws in recalculated.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value in errors
    ]
    assert not bad, f"formula errors in the export: {bad[:10]}"


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM is Windows-only")
def test_excel_formulas_agree_with_the_engine(recalculated, result):
    """Recalculate through Excel and compare every cell against Python.

    This is the only thing standing between the export and a silent divergence
    between the two implementations of the model.
    """
    if recalculated is None:
        pytest.skip("Excel not available on this machine")

    ws = recalculated[REGISTER]
    head = _header_row(ws)

    mismatches = []
    for i, b in enumerate(result.breakdowns):
        r = head + 1 + i

        for col, field in DIRECT.items():
            excel = ws.cell(row=r, column=col).value
            assert excel is not None, f"{b.name} column {col} did not calculate"
            delta = abs(Decimal(str(excel)) - getattr(b, field))
            if delta > TOLERANCE:
                mismatches.append(
                    f"{b.name} col{col} {field}: excel={excel} "
                    f"engine={getattr(b, field)} delta={delta}"
                )

        for col, field in NEGATED.items():
            excel = ws.cell(row=r, column=col).value
            assert excel is not None, f"{b.name} column {col} did not calculate"
            delta = abs(Decimal(str(excel)) + getattr(b, field))
            if delta > TOLERANCE:
                mismatches.append(
                    f"{b.name} col{col} {field}: excel={excel} "
                    f"engine=-{getattr(b, field)} delta={delta}"
                )

        for col, fn in DERIVED.items():
            excel = ws.cell(row=r, column=col).value
            assert excel is not None, f"{b.name} column {col} did not calculate"
            delta = abs(Decimal(str(excel)) - fn(b))
            if delta > TOLERANCE:
                mismatches.append(
                    f"{b.name} col{col}: excel={excel} engine={fn(b)} delta={delta}"
                )

    assert not mismatches, "\n".join(mismatches)


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM is Windows-only")
def test_register_arithmetic_closes(recalculated, result):
    """The register must be internally consistent, not merely match the engine.

    NET PAY has to equal what is actually handed over: gross, less every deduction.
    """
    if recalculated is None:
        pytest.skip("Excel not available on this machine")

    ws = recalculated[REGISTER]
    head = _header_row(ws)

    for i, b in enumerate(result.breakdowns):
        r = head + 1 + i
        cell = lambda c: Decimal(str(ws.cell(row=r, column=c).value))

        # de minimis + incentive + basic = gross
        assert abs(cell(5) + cell(6) + cell(2) - cell(7)) <= TOLERANCE, b.name
        # non taxable + net taxable = gross taxable
        assert abs(cell(12) + cell(13) - cell(11)) <= TOLERANCE, b.name
        # net pay = gross plus the (negative) deductions and withholding
        expected = cell(7) + cell(8) + cell(9) + cell(10) + cell(14)
        assert abs(cell(15) - expected) <= TOLERANCE, b.name


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM is Windows-only")
def test_exported_total_matches_the_workbook_figure(recalculated, result):
    """The headline number, computed by Excel from the exported formulas."""
    if recalculated is None:
        pytest.skip("Excel not available on this machine")

    ws = recalculated[REGISTER]
    head = _header_row(ws)
    total_row = head + 1 + len(result.breakdowns)

    total = Decimal(str(ws.cell(row=total_row, column=27).value))
    assert abs(total - result.totals.tax_saved_annual) <= TOLERANCE
    assert abs(total - Decimal("119507.00")) <= Decimal("0.01")
